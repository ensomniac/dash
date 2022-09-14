#!/usr/bin/python
#
# Ensomniac 2022 Ryan Martin, ryan@ensomniac.com
#                Andrew Stet, stetandrew@gmail.com

import os
import sys

from Dash.Users import Users
from csv import writer, reader
from collections import OrderedDict
from dateutil.parser import isoparse
from Dash.Utils import GetRandomID, FormatTime


# This is old and was never fully written out, but Altona uses it to a minor capacity
class CSV:
    def __init__(self, csv_root="", dash_context={}, all_data={}, exclude_keys=[], file=None):
        """
        Handle all non-gsheet CSV-related actions, such as exporting and importing.

        :param str csv_root: Root where csv file will be stored or retrieved from (default="")
        :param dict dash_context: Dash Context object (default={})
        :param dict all_data: (When exporting) Data object matching the structure of Collection.GetAll()["data"] (default={})
        :param list exclude_keys: Data keys to exclude when translating data to/from CSV (default=[])
        :param byte file: (When importing) File byte data from a file that was uploaded (default=None)
        """

        self.file = file
        self.csv_root = csv_root
        self.dash_context = dash_context
        self.exclude_keys = exclude_keys
        self.users_mod = Users(request_params={}, dash_context=self.dash_context)

        # Paths
        self.csv_export_root = os.path.join(self.csv_root, "export")
        self.csv_import_root = os.path.join(self.csv_root, "import")
        self.csv_export_filepath = os.path.join(self.csv_export_root, f"{GetRandomID()}.csv")
        self.csv_import_filepath = os.path.join(self.csv_import_root, f"{GetRandomID()}.csv")

        # Keys to move to the beginning and end, respectively, when sorting
        self.prioritized_keys = ["display_name"]
        self.unprioritized_keys = ["id", "created_by", "created_on", "modified_by", "modified_on"]

        if all_data:
            self.all_data = self.get_all_cleaned_data(all_data)
        else:
            self.all_data = all_data

        if self.csv_root:
            os.makedirs(self.csv_root, exist_ok=True)

    def GetRowData(self, csv_path, has_header=True):
        rows = []
        header = None
        csv_reader = reader(open(csv_path, "r"))

        if has_header:
            header = next(csv_reader)

        for row in csv_reader:
            if header:
                rows.append({header[index]: value for (index, value) in enumerate(row)})
            else:
                rows.append(row)

        return rows

    def ExtractCSVsFromExcel(self, excel_file_path, csv_filename="", filename_getter=None, sheet_name_filters={}, local=False):
        from openpyxl import load_workbook
        from Dash.Utils import GetAssetPath
        from Dash.LocalStorage import ConformPermissions

        if local:
            print("\nExtracting individual sheets from Excel file as CSVs...\n\tReading Excel file...")

        workbook = load_workbook(excel_file_path)

        for sheet_name in workbook.sheetnames:

            # Follow this pattern to expand sheet_name filtering
            if sheet_name_filters.get("startswith") and type(sheet_name_filters["startswith"]) is list:
                skip = False

                for name in sheet_name_filters["startswith"]:
                    if not sheet_name.startswith(name):
                        skip = True

                        break

                if skip:
                    continue

            if local:
                print(f"\tExtracting {sheet_name}...")

            sheet = workbook[sheet_name]

            if not csv_filename and filename_getter:
                filename = filename_getter(sheet_name)
            else:
                filename = csv_filename

                if filename.endswith(".csv"):
                    filename = filename.rstrip(".csv")

                filename = GetAssetPath(filename)

            if not filename.endswith(".csv"):
                filename += ".csv"

            csv_path = os.path.join(self.csv_root, filename)

            print(f"\t\tWriting to {csv_path}...")

            with open(csv_path, "w", newline="\n") as file_handle:
                csv_writer = writer(file_handle)

                for row in sheet.iter_rows():
                    csv_writer.writerow([cell.value for cell in row])
                    
            ConformPermissions(csv_path)

    def Export(self):
        os.makedirs(self.csv_export_root, exist_ok=True)

        header = False
        csv_file = open(self.csv_export_filepath, "w")
        csv_writer = writer(csv_file)

        for data in self.all_data:
            if not header:
                csv_writer.writerow(self.get_keys(data))
                header = True

            csv_writer.writerow(self.get_values(data))

        csv_file.close()

        return self.csv_export_filepath

    def Import(self):
        os.makedirs(self.csv_import_root, exist_ok=True)

        """
        This should only need to handle actually importing the file.
        Data management after the fact should be handled outside this script.

        :return: Imported CSV path.
        :rtype: str
        """

        try:
            open(self.csv_import_filepath, "wb").write(self.file)

            return self.csv_import_filepath
        except:
            return ""

    def ensure_roots_exist(self):
        for root in [self.csv_root, self.csv_export_root, self.csv_import_root]:
            os.makedirs(root, exist_ok=True)

    def get_all_cleaned_data(self, all_data):
        """
        Generic cleaning of data for proper data distribution over the CSV,
        as well as for improving user readability, like formatting timestamps, etc

        :param dict all_data: all data objects to clean
        :return: all cleaned data objects
        :rtype: list
        """

        all_keys = []
        all_cleaned_data = []
        sort_keys = ["display_name", "asset_path"]

        for item in all_data:
            data = all_data[item]

            for sort_key in sort_keys:
                if not data.get(sort_key):
                    continue

            data = self.replace_user_email_with_name(data)
            data = self.replace_timestamp_with_readable(data)

            for exclude_key in self.exclude_keys:
                if exclude_key in data:
                    del data[exclude_key]

            for key in data:
                if key not in all_keys:
                    all_keys.append(key)

            all_cleaned_data.append(data)

        for index, data in enumerate(all_cleaned_data):
            for key in all_keys:
                if key not in data:
                    data[key] = None

            all_cleaned_data[index] = self.reorder_data(data)

        for sort_key in sort_keys:
            if sort_key in all_keys:
                return self.sort_data_by_key(all_cleaned_data, sort_key)

        return all_cleaned_data

    def replace_user_email_with_name(self, data):
        for key in ["created_by", "modified_by"]:
            if not data.get(key):
                continue

            user_data = self.users_mod.GetUserData(data[key])

            if not user_data.get("first_name") and not user_data.get("last_name"):
                continue

            data[key] = f"{user_data['first_name']} {user_data['last_name']}"

        return data

    def replace_timestamp_with_readable(self, data):
        for key in ["created_on", "modified_on"]:
            if not data.get(key):
                continue

            data[key] = FormatTime(isoparse(data[key]))

        return data

    def sort_data_by_key(self, all_data, sort_key):
        """
        Sort data using designated key

        :param list all_data: json data dicts
        :param str sort_key: key to use when sorting
        :rtype: list
        :return: sorted list of json data dicts
        """

        sorted_data = []
        cleaned_data = {}

        for data in all_data:
            cleaned_data[data[sort_key]] = data

        cleaned_data_sorted = cleaned_data.items()

        try:
            cleaned_data_sorted = sorted(cleaned_data_sorted)
        except TypeError:
            pass

        cleaned_data = OrderedDict(cleaned_data_sorted)

        for key in cleaned_data:
            sorted_data.append(cleaned_data[key])

        return sorted_data

    def reorder_data(self, data):
        ordered_data = OrderedDict(sorted(data.items()))
        order = [item for item in ordered_data]

        for key in self.prioritized_keys:
            if key in order:
                order.remove(key)
                order.insert(0, key)

        for key in self.unprioritized_keys:
            if key in order:
                order.remove(key)
                order.append(key)

        for key in order:
            value = ordered_data[key]
            del ordered_data[key]
            ordered_data[key] = value

        return ordered_data

    def get_values(self, data):
        values = list(data.values())

        for index, value in enumerate(values):
            if type(value) is list:
                try:
                    values[index] = "\n".join(value)
                except:
                    values[index] = str(value)

        return values

    def get_keys(self, data):
        keys = list(data.keys())

        for index, key in enumerate(keys):
            if key == "id":
                keys[index] = "ID"

            elif key == "display_name":
                keys[index] = "Name"

            elif key == "phones":
                keys[index] = "Phone Numbers"

            elif key == "emails":
                keys[index] = "Email Addresses"

            elif key == "associated_orgs":
                keys[index] = "Associated Organizations"

            else:
                keys[index] = key.replace("_", " ").title()

        return keys


def GetRowData(csv_path, has_header=True):
    return CSV().GetRowData(csv_path, has_header)
