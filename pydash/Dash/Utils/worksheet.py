#!/usr/bin/python
#
# Ensomniac 2024 Ryan Martin, ryan@ensomniac.com
#             Andrew Stet, stetandrew@gmail.com

import os
import sys


class WorksheetUtils:
    def __init__(self, worksheet):
        self.worksheet = worksheet

        self.fonts = {}
        self.fills = {}
        self.sides = {}
        self.borders = {}
        self.alignments = {}

    def AutoSizeColumnsByContent(self, pad=0):
        from openpyxl.utils import get_column_letter

        # Auto-size columns based on content
        for column in self.worksheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value]

            for cell in column:
                try:
                    max_length = max(max_length, *map(len, str(cell.value).split("\n")))
                except:
                    pass

            self.worksheet.column_dimensions[get_column_letter(column[0].column)].width = (
                max_length + pad  # Adding a little extra space
            )

    def CenterTextVerticallyInAllCells(self, min_height=30, min_height_mults={}):
        alignment = self.get_alignment(vertical="center")

        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment

            if (
                min_height
                and (
                    row[0].row not in self.worksheet.row_dimensions
                    or self.worksheet.row_dimensions[row[0].row].height < min_height
                )
            ):
                self.worksheet.row_dimensions[row[0].row].height = min_height * (
                    min_height_mults[row[0].row] if row[0].row in min_height_mults else 1
                )

    def SetCellValueByColumnIndex(self, row_num, column_index, value=""):
        for col_index, cell in enumerate(self.worksheet[row_num]):
            if col_index != column_index:
                continue

            cell.value = value

            break

    def StyleHeaderRow(self, bg_color="dcdfe3", font=None, border=None, fill=None, font_color=""):
        font, border, fill = self.assert_header_footer_params(
            bg_color, font, border, fill, font_color, header=True
        )

        self.StyleRow(1, font, border, fill)

    def StyleFooterRow(self, bg_color="dcdfe3", font=None, border=None, fill=None, font_color=""):
        font, border, fill = self.assert_header_footer_params(bg_color, font, border, fill, font_color)

        self.StyleRow(self.worksheet.max_row, font, border, fill)

    def StyleCell(
        self, row_num, col_num, font=None, border=None, fill=None,
        bg_color="", font_type="", font_color="", include_border=True,
        alignment=None, align_hor="center", align_ver="center"
    ):
        return self.StyleRow(
            row_num=row_num,
            font=font,
            border=border,
            fill=fill,
            bg_color=bg_color,
            font_type=font_type,
            font_color=font_color,
            include_border=include_border,
            alignment=alignment,
            align_hor=align_hor,
            align_ver=align_ver,
            _col_num=col_num
        )

    def StyleRow(
        self, row_num, font=None, border=None, fill=None, bg_color="",
        font_type="", font_color="", include_border=True, alignment=None,
        align_hor="center", align_ver="center", _col_num=0
    ):
        fill, font, border, alignment = self.init_style_params(
            fill, bg_color, font, font_type, font_color, border, include_border, alignment, align_hor, align_ver
        )

        for index, cell in enumerate(self.worksheet[row_num]):
            if _col_num:
                i = index + 1

                if i < _col_num:
                    continue

                if i > _col_num:
                    break

            self.style_cell(cell, fill, font, border, alignment)

    def StyleColumn(
        self, col_letter_or_num, font=None, border=None, fill=None,
        bg_color="", font_type="", font_color="", include_border=True,
        alignment=None, align_hor="center", align_ver="center"
    ):
        if type(col_letter_or_num) is int:
            from openpyxl.utils import get_column_letter

            col_letter_or_num = get_column_letter(col_letter_or_num)

        if type(col_letter_or_num) is not str:
            raise ValueError(f"'col_letter_or_num' must be a letter or a number: {col_letter_or_num}")

        from openpyxl.utils import column_index_from_string

        col_index = column_index_from_string(col_letter_or_num.upper())

        fill, font, border, alignment = self.init_style_params(
            fill, bg_color, font, font_type, font_color, border, include_border, alignment, align_hor, align_ver
        )

        # Apply styles to each cell in the specified column
        for row in self.worksheet.iter_rows(min_col=col_index, max_col=col_index):
            for cell in row:
                self.style_cell(cell, fill, font, border, alignment)

    def get_font(self, font_type="", font_color=""):
        key = f"{font_type}{font_color}"

        if not self.fonts.get(key):
            from openpyxl.styles import Font

            kwargs = {}

            if font_color:
                kwargs["color"] = font_color

            if font_type:
                kwargs[font_type] = True

            self.fonts[key] = Font(**kwargs)

        return self.fonts[key]

    def get_fill(self, start_color, end_color="", fill_type="solid"):
        key = f"{start_color}{end_color}{fill_type}"

        if not self.fills.get(key):
            from openpyxl.styles import PatternFill

            self.fills[key] = PatternFill(
                start_color=start_color,
                end_color=end_color or start_color,
                fill_type=fill_type
            )

        return self.fills[key]

    def get_side(self, style="thin"):
        if not self.sides.get(style):
            from openpyxl.styles import Side

            self.sides[style] = Side(style=style)

        return self.sides[style]

    def get_alignment(self, horizontal="", vertical=""):
        key = f"{horizontal}{vertical}"

        if not self.alignments.get(key):
            from openpyxl.styles import Alignment

            self.alignments[key] = Alignment(
                horizontal=horizontal or None,
                vertical=vertical or None
            )

        return self.alignments[key]

    def get_border(self, left=None, right=None, top=None, bottom=None):
        key = f"{left}{right}{top}{bottom}"

        if not self.borders.get(key):
            from openpyxl.styles import Border

            self.borders[key] = Border(
                left=self.get_side(left) if left else None,
                right=self.get_side(right) if right else None,
                top=self.get_side(top) if top else None,
                bottom=self.get_side(bottom) if bottom else None
            )

        return self.borders[key]

    def assert_header_footer_params(
        self, bg_color="dcdfe3", font=None, border=None, fill=None, font_color="", header=False
    ):
        if not fill:
            fill = self.get_fill(bg_color)

        if not font:
            font = self.get_font("bold", font_color)

        if not border:
            border = self.get_border(**{"bottom" if header else "top": "thin"})

        return font, border, fill

    def init_style_params(
        self, fill, bg_color, font, font_type, font_color, border, include_border, alignment, align_hor, align_ver
    ):
        if not fill and bg_color:
            fill = self.get_fill(bg_color)

        if not font and (font_type or font_color):
            font = self.get_font(font_type, font_color)

        if not border and include_border:
            border = self.get_border("thin", "thin", "thin", "thin")

        if not alignment and (align_hor != "center" or align_ver != "center"):
            alignment = self.get_alignment(
                horizontal=align_hor,
                vertical=align_ver
            )

        return fill, font, border, alignment

    def style_cell(self, cell, fill, font, border, alignment):
        if fill:
            cell.fill = fill

        if font:
            cell.font = font

        if border:
            cell.border = border

        if alignment:
            cell.alignment = alignment
